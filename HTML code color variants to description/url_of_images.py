import requests
from bs4 import BeautifulSoup
import openpyxl

# Načíst Excel soubor
excel_file = "get_images.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Procházet URL adresy ve sloupci A
for row in range(2, sheet.max_row + 1):  # Předpokládáme, že první řádek obsahuje záhlaví
    url = sheet.cell(row=row, column=1).value
    if url:
        try:
            # Stáhnout HTML obsah stránky
            response = requests.get(url)
            response.raise_for_status()
            
            # Parsovat HTML kód
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Najít tag <a> s odpovídající třídou
            link_tag = soup.find("a", class_="p-main-image cloud-zoom")
            if not link_tag:  # Pokud nebyl nalezen, hledáme druhou možnost
                link_tag = soup.find("a", class_="p-main-image cloud-zoom cbox")
            
            # Pokud byl nalezen, extrahujeme href
            if link_tag and "href" in link_tag.attrs:
                image_url = link_tag["href"]
                sheet.cell(row=row, column=2).value = image_url  # Zapsat URL obrázku do sloupce B
            else:
                sheet.cell(row=row, column=2).value = "Obrázek nenalezen"
        except Exception as e:
            sheet.cell(row=row, column=2).value = f"Chyba: {str(e)}"

# Uložit změny do Excel souboru
workbook.save(excel_file)
print("Hotovo! URL adresy obrázků byly přidány do sloupce B.")
