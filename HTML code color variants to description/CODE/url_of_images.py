import openpyxl
from bs4 import BeautifulSoup
import requests

# Načtení Excel souboru
excel_file = 'vase_soubor.xlsx'  # Změňte na název vašeho Excel souboru
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Procházení buněk ve sloupci B
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
    for cell in row:
        url = cell.value
        try:
            # Získání obsahu URL
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Vyhledání URL obrázku
            img_tag = soup.find("a", class_="p-main-image cloud-zoom") or soup.find("a", class_="p-main-image cloud-zoom cbox")
            img_url = img_tag['href'] if img_tag else 'Obrázek nenalezen'
            
            # Zapsání URL obrázku do sloupce C
            sheet.cell(row=cell.row, column=3).value = img_url
        except Exception as e:
            # Zapsání chybové zprávy
            sheet.cell(row=cell.row, column=3).value = f'Chyba: {str(e)}'

# Uložení změn do Excel souboru
wb.save(excel_file)
print("Hotovo!")
