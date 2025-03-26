from openpyxl import load_workbook

# Načtení existujícího Excel souboru
file_name = "shoptet_product.xlsx"
wb = load_workbook(file_name)
ws = wb.active

# Získání nadpisu od uživatele
title = input("Zadejte nadpis: ")

# Přidání nadpisu do pátého sloupce
ws["E1"] = f'''<div class="bik_htitle">
<h4>{title}</h4>
</div>
<div class="cyp_Galery-container"><!-- Šipky pro mobilní zobrazení -->
<div class="scroll-arrows right">&nbsp;</div>
<!-- Produkty -->'''

# Vložení HTML funkce do pátého sloupce pro zbytek řádků
for row in ws.iter_rows(min_row=2, values_only=False):  # Zpracovává řádky od 2
    a_value = row[0].value  # Hodnota ze sloupce A
    b_value = row[1].value  # Hodnota ze sloupce B
    c_value = row[2].value  # Hodnota ze sloupce C

    if a_value and b_value and c_value:  # Pokud všechny hodnoty existují
        html_content = f'''
        <div class="cyp_Galery-box">
        <div class="cyp_product-name"><span class="BIK_variant-title">{a_value}</span></div>
        <a href="{b_value}"><img loading="lazy" src="{c_value}" alt="{a_value}" /></a></div>
        '''
        ws[f"E{row[0].row}"] = html_content

# Přidání nadpisu do šestého sloupce
ws["F1"] = "</div></div></div></div>"
# Pomocný seznam pro porovnání unikátních hodnot
data = []

# Procházení dat a výpočet hodnot pro sloupec F
for row in ws.iter_rows(min_row=2, values_only=False):  # Zpracovává řádky od 2
    d_value = row[3].value  # Hodnota ze sloupce D
    e_value = row[4].value  # Hodnota ze sloupce E

    if d_value and e_value:  # Pokud hodnoty v D a E existují
        # Vytvoření kombinovaného textu na základě unikátních hodnot
        combined_values = []
        for comparison_row in ws.iter_rows(min_row=2, values_only=False):
            d_comp = comparison_row[3].value
            e_comp = comparison_row[4].value
            if d_comp == d_value and e_comp != e_value and e_comp not in combined_values:
                combined_values.append(e_comp)

        # Spojení hodnot pomocí funkce podobné TEXTJOIN
        textjoin_result = e_value + "".join(combined_values)
        ws[f"F{row[0].row}"] = textjoin_result
    else:
        # Pokud hodnoty chybí, vložme do F původní hodnotu z E
        ws[f"F{row[0].row}"] = e_value

# Přidání nadpisu do dalšího sloupce
ws["G1"] = "With URL"

# Procházení řádků a generování hodnot
for row in ws.iter_rows(min_row=2, values_only=False):  # Zpracovává od druhého řádku
    e_value = row[4].value  # Hodnota ze sloupce E
    f_value = row[5].value  # Hodnota ze sloupce F
    e1_value = ws["E1"].value  # Hodnota z buňky E1
    f1_value = ws["F1"].value  # Hodnota z buňky F1

    # Sestavení výsledného textu
    if e_value and f_value and e1_value and f1_value:
        result = (
            '<div id="colour_variants">'
            + e1_value
            + (f_value if f_value else e_value)
            + f1_value
            + "</div> <!-- konec div #colour_variants -->"
        )
        row[6].value = result  # Vložení do sloupce G
    else:
        # Pokud některá hodnota chybí, vloží se jen základní struktura
        row[6].value = "<div id=\"colour_variants\">Incomplete data</div>"

# Přidání nadpisu do sloupce H
ws["H1"] = "HTML code"

# Procházení řádků a generování hodnot
for row in ws.iter_rows(min_row=2, values_only=False):  # Zpracovává řádky od druhého
    g_value = row[6].value  # Hodnota ze sloupce G
    b_value = row[1].value  # Hodnota ze sloupce B

    # Pokud existuje hodnota ve sloupci G a B, nahradíme část textu
    if g_value and b_value:
        html_code = g_value.replace(f'<a href="{b_value}"', '<a href="#"')
        row[7].value = html_code  # Vložení upraveného HTML kódu do sloupce H
    else:
        # Pokud některá z hodnot chybí, vloží informaci o neúplných datech
        row[7].value = "Incomplete data"

# Uložení výsledků zpět do Excelu
wb.save(file_name)
print(f"Hotové hodnoty byly spočítány a uloženy v souboru {file_name}.")

