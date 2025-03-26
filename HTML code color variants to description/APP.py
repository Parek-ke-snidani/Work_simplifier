import tkinter as tk
from tkinter import messagebox
import requests
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from tkinter import filedialog

def check_and_add_input(entry):
    """Check the value and add a new input field if the current field is filled."""
    if entry.get().strip() and not hasattr(entry, "next_added"):
        entry.next_added = True
        add_url_input()

def add_url_input():
    """Add a new input field for URLs and set up its monitoring."""
    entry = tk.Entry(frame_urls, width=50)
    entry.pack(pady=2)
    url_entries.append(entry)
    entry.bind("<KeyRelease>", lambda e: check_and_add_input(entry))

def start_scraping():
    """Start scraping using URLs entered by the user."""
    category_urls = [entry.get().strip() for entry in url_entries if entry.get().strip()]
    if not category_urls:
        messagebox.showerror("Error", "Please enter at least one URL.")
        return

    all_products = []
    for category in category_urls:
        max_pages = get_max_pages(category)
        print(f"🔍 Category: {category} (Max {max_pages} pages)")
        
        for page in range(1, max_pages + 1):
            url = f"{category}strana-{page}/"
            print(f"📄 Loading: {url}")
            all_products.extend(get_products(url))
    
    if all_products:
        save_to_excel(all_products, "shoptet_product.xlsx")
        messagebox.showinfo("Done", "Scraping completed! Data saved to shoptet_product.xlsx.")
    else:
        messagebox.showinfo("Done", "No data found.")

def save_to_excel(data, file_path):
    """Save data to an Excel file."""
    df = pd.DataFrame(data, columns=["name", "Product URL"])
    df = df.drop_duplicates()
    df.to_excel(file_path, index=False, engine="openpyxl")
    print(f"✅ Data saved to {file_path}")

def check_url_status():
    """Check the status of URLs in the Excel file."""
    file_path = "shoptet_product.xlsx"
    try:
        df = pd.read_excel(file_path)
    except FileNotFoundError:
        messagebox.showerror("Error", "The file shoptet_product.xlsx was not found.")
        return

    unavailable_count = 0
    for index, url in enumerate(df['Product URL']):
        try:
            response = requests.head(url, allow_redirects=True, timeout=5)
            if response.status_code < 400:
                df.loc[index, 'Status'] = "Available"
            else:
                df.loc[index, 'Status'] = f"Unavailable (Status code: {response.status_code})"
                unavailable_count += 1
        except requests.RequestException as e:
            df.loc[index, 'Status'] = f"Error ({e})"
            unavailable_count += 1

    df.to_excel(file_path, index=False, engine="openpyxl")
    messagebox.showinfo("URL Status", f"URL status check completed! Unavailable URLs: {unavailable_count}")

def fetch_images():
    """Fetch image URLs from product URLs in the Excel file."""
    file_path = "shoptet_product.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        messagebox.showerror("Error", "The file shoptet_product.xlsx was not found.")
        return

    missing_images_count = 0
    for row in range(2, sheet.max_row + 1):  # Skip header row
        url = sheet.cell(row=row, column=2).value  # Assumes product URLs are in column B
        if url:
            try:
                response = requests.get(url)
                response.raise_for_status()

                soup = BeautifulSoup(response.text, 'html.parser')
                link_tag = soup.find("a", class_="p-main-image cloud-zoom")
                if not link_tag:
                    link_tag = soup.find("a", class_="p-main-image cloud-zoom cbox")

                if link_tag and "href" in link_tag.attrs:
                    image_url = link_tag["href"]
                    sheet.cell(row=row, column=3).value = image_url  # Save image URL to column C
                else:
                    sheet.cell(row=row, column=3).value = "Image not found"
                    missing_images_count += 1
            except Exception as e:
                sheet.cell(row=row, column=3).value = f"Error: {str(e)}"
                missing_images_count += 1

    workbook.save(file_path)
    messagebox.showinfo("Image Extraction", f"Image extraction completed! Missing images: {missing_images_count}")

def get_max_pages(category_url):
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(category_url, headers=headers)
    response.encoding = response.apparent_encoding  

    if response.status_code != 200:
        print(f"❌ Cannot load {category_url}")
        return 1  

    soup = BeautifulSoup(response.text, "html.parser")
    last_page_link = soup.find("a", {"data-testid": "linkLastPage"})
    if last_page_link:
        try:
            return int(last_page_link.get("href").split("-")[-1].replace("/", ""))
        except ValueError:
            return 1

    return 1

def get_products(category_url):
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(category_url, headers=headers)
    response.encoding = response.apparent_encoding  

    if response.status_code != 200:
        print(f"❌ Error loading {category_url}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")
    products = []
    
    for product in soup.find_all("div", {"data-micro": "product"}):
        name_tag = product.find("span", {"data-micro": "name"})
        name = name_tag.text.strip() if name_tag else "N/A"

        url_tag = product.find("a", {"data-micro": "url"})
        url = "N/A"
        if url_tag:
            url = url_tag.get("href")
            if url.startswith("/"):
                url = "https://www.bikemax.cz" + url

        products.append((name, url))

    return products

def process_excel_data():
    """Process the Excel file based on the given instructions."""
    file_name = "shoptet_product.xlsx"
    try:
        wb = load_workbook(file_name)
        ws = wb.active

        # Add header to the fourth column if it's not filled
        if not ws["D1"].value:
            ws["D1"] = "Made up code"

        # Iterate through rows and calculate values from formulas
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):  # Column A (index 0)
            cell_value = row[0].value  # Value from column A
            try:
                # Attempt to remove text between parentheses
                if cell_value and "(" in cell_value and ")" in cell_value:
                    start_index = cell_value.find(" (")
                    end_index = cell_value.find(")", start_index)
                    if start_index != -1 and end_index != -1:
                        new_value = cell_value[:start_index] + cell_value[end_index + 1:]
                        ws[f"D{row[0].row}"] = new_value
                else:
                    # Copy original value if parentheses are not found
                    ws[f"D{row[0].row}"] = cell_value
            except Exception as e:
                # In case of any error, retain original value from column A
                ws[f"D{row[0].row}"] = cell_value

        # Save the updated workbook
        wb.save(file_name)
        messagebox.showinfo("Excel Processing", "Excel file updated successfully!")
    except FileNotFoundError:
        messagebox.showerror("Error", "The file shoptet_product.xlsx was not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def generate_html_code():
    """Generate HTML content and update Excel file based on provided logic."""
    from openpyxl import load_workbook
    from tkinter.simpledialog import askstring
    from tkinter import messagebox

    file_name = "shoptet_product.xlsx"
    try:
        wb = load_workbook(file_name)
        ws = wb.active

        # Získání nadpisu od uživatele pomocí vyskakovacího okna
        title = askstring("Zadejte nadpis", "Zadejte nadpis pro generování HTML:")
        if not title:
            messagebox.showerror("Chyba", "Nadpis nebyl zadán. Akce byla zrušena.")
            return

        # Přidání nadpisu do pátého sloupce
        ws["E1"] = f'''<div class="bik_htitle">
        <h4>{title}</h4>
        </div>
        <div class="cyp_Galery-container"><!-- Šipky pro mobilní zobrazení -->
        <div class="scroll-arrows right">&nbsp;</div>
        <!-- Produkty -->'''

        # Vložení HTML obsahu do pátého sloupce pro zbytek řádků
        for row in ws.iter_rows(min_row=2, values_only=False):  # Zpracovává od druhého řádku
            a_value = row[0].value
            b_value = row[1].value
            c_value = row[2].value

            if a_value and b_value and c_value:
                html_content = f'''
                <div class="cyp_Galery-box">
                <div class="cyp_product-name"><span class="BIK_variant-title">{a_value}</span></div>
                <a href="{b_value}"><img loading="lazy" src="{c_value}" alt="{a_value}" /></a></div>
                '''
                ws[f"E{row[0].row}"] = html_content

        # Přidání nadpisu do šestého sloupce
        ws["F1"] = "</div></div></div></div>"

        # Zpracování unikátních dat ve sloupci F
        for row in ws.iter_rows(min_row=2, values_only=False):  # Zpracovává od druhého řádku
            d_value = row[3].value
            e_value = row[4].value

            if d_value and e_value:
                combined_values = []
                for comparison_row in ws.iter_rows(min_row=2, values_only=False):
                    d_comp = comparison_row[3].value
                    e_comp = comparison_row[4].value
                    if d_comp == d_value and e_comp != e_value and e_comp not in combined_values:
                        combined_values.append(e_comp)

                textjoin_result = e_value + "".join(combined_values)
                ws[f"F{row[0].row}"] = textjoin_result
            else:
                ws[f"F{row[0].row}"] = e_value

        # Přidání nadpisu do sedmého sloupce
        ws["G1"] = "With URL"

        # Generování hodnot pro sedmý sloupec
        for row in ws.iter_rows(min_row=2, values_only=False):
            e_value = row[4].value
            f_value = row[5].value
            e1_value = ws["E1"].value
            f1_value = ws["F1"].value

            if e_value and f_value and e1_value and f1_value:
                result = (
                    '<div id="colour_variants">'
                    + e1_value
                    + (f_value if f_value else e_value)
                    + f1_value
                    + "</div> <!-- konec div #colour_variants -->"
                )
                row[6].value = result
            else:
                row[6].value = "<div id=\"colour_variants\">Incomplete data</div>"

        # Přidání nadpisu do osmého sloupce
        ws["H1"] = "HTML code"

        # Generování hodnot pro osmý sloupec
        for row in ws.iter_rows(min_row=2, values_only=False):
            g_value = row[6].value
            b_value = row[1].value

            if g_value and b_value:
                html_code = g_value.replace(f'<a href="{b_value}"', '<a href="#"')
                row[7].value = html_code
            else:
                row[7].value = "Incomplete data"

        # Uložení aktualizovaného souboru
        wb.save(file_name)
        messagebox.showinfo("Hotovo", f"HTML generování bylo dokončeno a uloženo v souboru {file_name}.")
    except FileNotFoundError:
        messagebox.showerror("Chyba", "Excel soubor nebyl nalezen!")
    except Exception as e:
        messagebox.showerror("Chyba", f"Došlo k chybě: {str(e)}")

def open_process_menu():
    """Open a new menu with full functionality of the provided code."""
    menu_window = tk.Toplevel(root)
    menu_window.title("Excel Processor")
    menu_window.geometry("500x400")

    # Initialize file paths
    exported_file_path = ""
    template_file_path = ""

    def select_exported_file():
        """Allow user to select the exported file."""
        nonlocal exported_file_path
        exported_file_path = filedialog.askopenfilename(title="Vyberte exported file")
        exported_label.config(text=f"Exported file: {exported_file_path}")

    def select_template_file():
        """Allow user to select the template file."""
        nonlocal template_file_path
        template_file_path = filedialog.askopenfilename(title="Vyberte Template file")
        template_label.config(text=f"Template file: {template_file_path}")

    def process_files():
        """Process the exported and template files."""
        if not exported_file_path or not template_file_path:
            output_label.config(text="Prosím, vyberte oba soubory.")
            return

        try:
            # Load Excel files into pandas DataFrames
            exported_df = pd.read_excel(exported_file_path)
            template_df = pd.read_excel(template_file_path)

            # Data processing based on columns "name" and "HTML code"
            for index, row in exported_df.iterrows():
                name_value = row["name"]
                description = row["description"]

                # Replace NaN description with an empty string
                if pd.isna(description):
                    description = ""

                # Find corresponding "HTML code" value in the template file
                html_code_row = template_df[template_df["name"] == name_value]
                if not html_code_row.empty:
                    html_code = html_code_row.iloc[0]["HTML code"]

                    # Check description content and append HTML code if necessary
                    if "<div id=\"colour_variants\">" not in description:
                        exported_df.at[index, "description"] = description + html_code

            # Automatically save the modified file as "for_import.xlsx"
            output_file_path = "for_import.xlsx"
            exported_df.to_excel(output_file_path, index=False)
            output_label.config(text=f"Soubor byl uložen jako '{output_file_path}'!")
        except Exception as e:
            output_label.config(text=f"Došlo k chybě: {str(e)}")

    # GUI elements for the new menu
    intro_label = tk.Label(menu_window, text="choose exported file a Template file:")
    intro_label.pack()

    exported_label = tk.Label(menu_window, text="Exported file: [none]")
    exported_label.pack()

    exported_button = tk.Button(menu_window, text="choose Exported file", command=select_exported_file)
    exported_button.pack()

    template_label = tk.Label(menu_window, text="Template file: [none]")
    template_label.pack()

    template_button = tk.Button(menu_window, text="choose Template file", command=select_template_file)
    template_button.pack()

    process_button = tk.Button(menu_window, text="Generate file", command=process_files)
    process_button.pack()

    output_label = tk.Label(menu_window, text="")
    output_label.pack()

# GUI application
root = tk.Tk()
root.title("Scraper GUI")

url_entries = []

frame_urls = tk.Frame(root)
frame_urls.pack(pady=10)

btn_start_scraping = tk.Button(root, text="Get URLs from categories", command=start_scraping)
btn_start_scraping.pack(pady=5)

btn_check_urls = tk.Button(root, text="Check URLs Status", command=check_url_status)
btn_check_urls.pack(pady=5)

btn_fetch_images = tk.Button(root, text="Get Images from URLs", command=fetch_images)
btn_fetch_images.pack(pady=5)

btn_process_excel = tk.Button(root, text="Generate Pair codes", command=process_excel_data)
btn_process_excel.pack(pady=5)

btn_generate_html = tk.Button(root, text="Generate HTML Code", command=generate_html_code)
btn_generate_html.pack(pady=5)

btn_open_menu = tk.Button(root, text="Generate import file", command=open_process_menu)
btn_open_menu.pack(pady=5)

add_url_input()  # Add the first input field

root.mainloop()
