from openpyxl import load_workbook

# Načtení existujícího Excel souboru
file_name = "shoptet_product.xlsx"
wb = load_workbook(file_name)
ws = wb.active

# Přidání nadpisu do čtvrtého sloupce, pokud není vyplněn
if not ws["D1"].value:
    ws["D1"] = "Made up code"

# Procházení řádků a výpočet hodnot ze vzorců
for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):  # Sloupec A (index 0)
    cell_value = row[0].value  # Hodnota buňky ve sloupci A
    try:
        # Pokusíme se najít a odstranit část textu mezi závorkami
        if cell_value and "(" in cell_value and ")" in cell_value:
            start_index = cell_value.find(" (")
            end_index = cell_value.find(")", start_index)
            if start_index != -1 and end_index != -1:
                new_value = cell_value[:start_index] + cell_value[end_index + 1:]
                ws[f"D{row[0].row}"] = new_value
        else:
            # Pokud závorky nejsou nalezeny, zkopírujeme původní hodnotu
            ws[f"D{row[0].row}"] = cell_value
    except Exception as e:
        # Pokud dojde k jakékoli chybě, vloží původní hodnotu ze sloupce A
        ws[f"D{row[0].row}"] = cell_value